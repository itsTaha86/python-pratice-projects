from flask import Flask, render_template, redirect, url_for, request, jsonify
import uuid
import json
import os
from openpyxl import load_workbook

data_path = "to-do-web/data/notes.json"

app = Flask(__name__, template_folder="templates", static_folder="statics")


def load_notes():
    if os.path.exists(data_path):
        try:
            with open(data_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except json.JSONDecodeError:
            return []
    else:
        return []


notes = load_notes()


def save_notes(notes):
    os.makedirs("data", exist_ok=True)
    with open(data_path, "w", encoding="utf-8") as f:
        json.dump(notes, f, ensure_ascii=False, indent=4)


def load_excel(file):

    wb = load_workbook(file)
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]
    data_rows = rows[1:]

    notes_from_excel = []

    for row in data_rows:
        note = {
            "id": str(uuid.uuid4()),
            "title": row[0],
            "content": row[1],
            "done": row[2] if len(row) > 2 else False,
        }
        notes_from_excel.append(note)
    return notes_from_excel


@app.get("/")
def index():

    return render_template("index.html", notes=notes)


@app.post("/create")
def create_note():

    app.logger.debug(request.form.get("content"))

    notes.append(
        {
            "id": str(uuid.uuid4()),
            "title": request.form.get("title"),
            "content": request.form.get("content"),
            "done": False,
        }
    )
    save_notes(notes)
    return redirect(url_for("index"))


@app.post("/edit/<note_id>")
def edit_note(note_id):

    app.logger.debug(request.form.get("content"))

    for note in notes:

        if note["id"] == note_id:
            note["title"] = request.form.get("title")
            note["content"] = request.form.get("content")

    save_notes(notes)

    return redirect(url_for("index"))


@app.post("/delete/<note_id>")
def delete_note(note_id):

    global notes

    notes = [note for note in notes if note["id"] != note_id]

    save_notes(notes)

    return redirect(url_for("index"))


@app.post("/toggle/<note_id>")
def toggle_note(note_id):
    for note in notes:
        if note["id"] == note_id:
            note["done"] = not note["done"]

    save_notes(notes)

    return redirect(url_for("index"))


@app.post("/getxl")
def get_excel():
    file = request.files["file"]

    try:
        new_notes = load_excel(file)
    except Exception as e:
        return f"Error on reading Excel: {e}"

    notes.extend(new_notes)
    save_notes(notes)
    return redirect(url_for("index"))


@app.get("/getjson")
@app.get("/getjson")
def get_json():
    with open(data_path, "r", encoding="utf-8") as f:
        notes = json.load(f)
    return jsonify(notes)


if __name__ == "__main__":
    app.run(debug=True)
